from io import BytesIO
from typing import List, Dict, Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                v = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(v))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)


def generate_excel(trip_info: Dict[str, Any], expenses: List[Dict[str, Any]]) -> bytes:
    """
    Expected inputs:
      trip_info: dict built in app.py
      expenses: list of dicts with keys:
        category, expense_date, paid_by, description, amount, receipt_file (optional)
    Returns:
      Excel file as raw bytes
    """

    wb = Workbook()

    # -------------------------
    # Sheet 1: Summary
    # -------------------------
    ws = wb.active
    ws.title = "Summary"

    header_fill = PatternFill("solid", fgColor="EEEEEE")
    bold = Font(bold=True)

    ws["A1"] = "PERFORMA"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A1"].alignment = Alignment(horizontal="left")

    ws["A2"] = "Performa Expense Report"
    ws["A2"].font = Font(bold=True, size=12)
    ws["A2"].alignment = Alignment(horizontal="left")

    ws["A3"] = "© 2026 bkw"
    ws["A3"].font = Font(size=9, italic=True, color="666666")
    ws["A3"].alignment = Alignment(horizontal="left")

    ws["A5"] = "Employee Name"
    ws["B5"] = trip_info.get("employee_name", "")
    ws["A6"] = "Employee Email"
    ws["B6"] = trip_info.get("employee_email", "")
    ws["A7"] = "Trip Location"
    ws["B7"] = trip_info.get("location", "")
    ws["A8"] = "Business Purpose"
    ws["B8"] = trip_info.get("purpose", "")
    ws["A9"] = "Departure Date"
    ws["B9"] = trip_info.get("departure_date", "")
    ws["A10"] = "Return Date"
    ws["B10"] = trip_info.get("return_date", "")
    ws["A11"] = "Trip Days"
    ws["B11"] = trip_info.get("trip_days", 0)

    ws["A13"] = "Per Diem First/Last Day Rate"
    ws["B13"] = float(trip_info.get("per_diem_first_last_rate", 0) or 0)
    ws["A14"] = "Per Diem Middle Day Rate"
    ws["B14"] = float(trip_info.get("per_diem_middle_rate", 0) or 0)
    ws["A15"] = "Per Diem Days"
    ws["B15"] = trip_info.get("trip_days", 0)
    ws["A16"] = "Per Diem Total"
    ws["B16"] = float(trip_info.get("per_diem_total", 0) or 0)

    ws["A18"] = "Total Spend"
    ws["B18"] = float(trip_info.get("total_spend", 0) or 0)
    ws["A19"] = "Company Paid"
    ws["B19"] = float(trip_info.get("company_paid", 0) or 0)
    ws["A20"] = "Employee Paid"
    ws["B20"] = float(trip_info.get("employee_paid", 0) or 0)
    ws["A21"] = "Reimbursement Due"
    ws["B21"] = float(trip_info.get("reimbursement_due", 0) or 0)

    for r in range(5, 22):
        ws[f"A{r}"].font = bold

    for cell in ["B13", "B14", "B16", "B18", "B19", "B20", "B21"]:
        ws[cell].number_format = '"$"#,##0.00'
    ws["B15"].number_format = "0"

    ws["B8"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[8].height = 45

    _auto_width(ws)

    # -------------------------
    # Sheet 2: Line Items
    # -------------------------
    ws2 = wb.create_sheet("Line Items")

    headers = ["Category", "Expense Date", "Description", "Paid By", "Amount", "Receipt Attached"]
    ws2.append(headers)

    for c in range(1, len(headers) + 1):
        cell = ws2.cell(row=1, column=c)
        cell.font = bold
        cell.fill = header_fill

    for e in expenses or []:
        category = e.get("category", "")
        expense_date = e.get("expense_date", "")
        description = e.get("description", "")
        paid_by = e.get("paid_by", "")
        amount = float(e.get("amount", 0) or 0)
        receipt_attached = "Yes" if e.get("receipt_file") else "No"

        ws2.append([category, expense_date, description, paid_by, amount, receipt_attached])

    # Currency format for Amount column
    for row in range(2, ws2.max_row + 1):
        ws2.cell(row=row, column=5).number_format = '"$"#,##0.00'
        ws2.cell(row=row, column=3).alignment = Alignment(wrap_text=True)

    _auto_width(ws2)

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
