import base64
import os
import re
from datetime import date, timedelta
from io import BytesIO
from typing import List, Dict, Any, Optional, Tuple

import streamlit as st
from sendgrid import SendGridAPIClient
from sendgrid.helpers.mail import (
    Mail,
    Email,
    To,
    Cc,
    Attachment,
    FileContent,
    FileName,
    FileType,
    Disposition,
)

from openpyxl import load_workbook

from excel_generator import generate_excel
from per_diem import calculate_per_diem_total, format_per_diem_summary
from sendgrid_secrets import (
    format_sendgrid_error,
    has_sendgrid_key,
    make_sendgrid_b64,
    secret_float,
    secret_get,
    sendgrid_api_key,
    sendgrid_key_diagnostics,
    verify_sendgrid_key_with_api,
)


# -----------------------------
# Config and constants
# -----------------------------
st.set_page_config(page_title="Performa Expense Report", layout="wide")

# Guam per diem: $75 first/last day, $100 middle days (overridable via secrets)
PER_DIEM_MIDDLE_RATE = secret_float(
    "PER_DIEM_MIDDLE_RATE", secret_float("PER_DIEM_RATE", 100.0)
)
PER_DIEM_FIRST_LAST_RATE = secret_float("PER_DIEM_FIRST_LAST_RATE", 75.0)

# NN per diem rates
NN_TRAVEL_DAY_RATE = secret_float("NN_TRAVEL_DAY_RATE", 51.0)
NN_NON_TRAVEL_DAY_RATE = secret_float("NN_NON_TRAVEL_DAY_RATE", 68.0)

MAX_ATTACHMENT_MB = secret_float("MAX_ATTACHMENT_MB", 18.0)

TEMPLATES_DIR = "templates"
NN_TEMPLATE_FILENAME = "nn_templates.xlsx"
NN_SHEET_NAME_DEFAULT = "Traveler Expense"

# NN daily grid settings
NN_START_ROW = 3
NN_COL_DATE = 1          # A
NN_COL_BREAKFAST = 2     # B (do not change)
NN_COL_LUNCH = 3         # C (do not change)
NN_COL_DINNER = 4        # D (do not change)
NN_COL_INCIDENTALS = 5   # E (we set per diem here)
NN_COL_AIRFARE = 6       # F
NN_COL_LODGING = 7       # G
NN_COL_VEHICLE_MILEAGE = 8  # H (not used)
NN_COL_RENTAL_CAR = 9    # I
NN_COL_MISC_TRAVEL = 10  # J
NN_COL_MIE_TOTALS = 11   # K (formula)
NN_COL_MIE_MAX = 12      # L (formula)
NN_COL_NOTES = 13        # M

CATEGORIES = [
    "Airfare",
    "Airport Parking",
    "Taxi or Uber to Airport",
    "Hotel",
    "Rental Car",
    "Gas for Rental Car",
    "Other",
]


# -----------------------------
# Helpers
# -----------------------------
def missing_email_secrets() -> List[str]:
    missing = []
    if not has_sendgrid_key():
        missing.append("SENDGRID_API_KEY_PART1 + PART2 (one line each; use PART3/PART4 if needed)")
    for key in ("SENDER_EMAIL", "FINANCE_EMAIL", "APPROVER_EMAIL"):
        if not secret_get(key).strip():
            missing.append(key)
    return missing


def bytes_from_uploaded_file(uploaded_file) -> bytes:
    if uploaded_file is None:
        return b""
    return uploaded_file.getvalue()


def calc_trip_days(departure: date, ret: date) -> int:
    if not departure or not ret:
        return 0
    if ret < departure:
        return 0
    return (ret - departure).days + 1


def calc_totals(expenses: List[Dict[str, Any]]) -> Dict[str, float]:
    total_spend = 0.0
    company_paid = 0.0
    employee_paid = 0.0

    for e in expenses:
        amt = float(e.get("amount") or 0)
        total_spend += amt
        if e.get("paid_by") == "Performa":
            company_paid += amt
        else:
            employee_paid += amt

    return {
        "total_spend": total_spend,
        "company_paid": company_paid,
        "employee_paid": employee_paid,
    }


def build_email_html(
    employee_name: str,
    employee_email: str,
    location: str,
    purpose: str,
    departure_date: date,
    return_date: date,
    per_diem_total: float,
    total_spend: float,
    company_paid: float,
    employee_paid: float,
    reimbursement_due: float,
    expenses: List[Dict[str, Any]],
) -> str:
    # No em dashes used
    def esc(x: Optional[str]) -> str:
        if x is None:
            return ""
        return (
            str(x)
            .replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
        )

    lines_html = ""
    if expenses:
        rows = []
        for i, e in enumerate(expenses, start=1):
            rows.append(
                f"""
                <tr>
                  <td style="padding:6px 8px;border-bottom:1px solid #eee;">{i}</td>
                  <td style="padding:6px 8px;border-bottom:1px solid #eee;">{esc(e.get("category",""))}</td>
                  <td style="padding:6px 8px;border-bottom:1px solid #eee;">{esc(e.get("expense_date",""))}</td>
                  <td style="padding:6px 8px;border-bottom:1px solid #eee;">{esc(e.get("description",""))}</td>
                  <td style="padding:6px 8px;border-bottom:1px solid #eee;">{esc(e.get("paid_by",""))}</td>
                  <td style="padding:6px 8px;border-bottom:1px solid #eee;text-align:right;">${float(e.get("amount") or 0):,.2f}</td>
                  <td style="padding:6px 8px;border-bottom:1px solid #eee;">{"Yes" if e.get("receipt_file") else "No"}</td>
                </tr>
                """
            )
        lines_html = f"""
        <p><strong>Line items:</strong></p>
        <table style="border-collapse:collapse;width:100%;font-family:Arial, sans-serif;font-size:13px;">
          <thead>
            <tr>
              <th style="text-align:left;padding:6px 8px;border-bottom:2px solid #ddd;">#</th>
              <th style="text-align:left;padding:6px 8px;border-bottom:2px solid #ddd;">Category</th>
              <th style="text-align:left;padding:6px 8px;border-bottom:2px solid #ddd;">Date</th>
              <th style="text-align:left;padding:6px 8px;border-bottom:2px solid #ddd;">Description</th>
              <th style="text-align:left;padding:6px 8px;border-bottom:2px solid #ddd;">Paid By</th>
              <th style="text-align:right;padding:6px 8px;border-bottom:2px solid #ddd;">Amount</th>
              <th style="text-align:left;padding:6px 8px;border-bottom:2px solid #ddd;">Receipt</th>
            </tr>
          </thead>
          <tbody>
            {''.join(rows)}
          </tbody>
        </table>
        """

    html = f"""
    <div style="font-family:Arial, sans-serif;font-size:14px;color:#111;">
      <p>Dear Performa Finance,</p>

      <p>Please find attached the submitted expense report for <strong>{esc(employee_name)}</strong> and accompanying receipts.</p>

      <p><strong>Details below:</strong></p>

      <table style="border-collapse:collapse;font-family:Arial, sans-serif;font-size:13px;">
        <tr><td style="padding:4px 10px 4px 0;"><strong>Employee Name:</strong></td><td style="padding:4px 0;">{esc(employee_name)}</td></tr>
        <tr><td style="padding:4px 10px 4px 0;"><strong>Employee Email:</strong></td><td style="padding:4px 0;">{esc(employee_email)}</td></tr>
        <tr><td style="padding:4px 10px 4px 0;"><strong>Trip Location:</strong></td><td style="padding:4px 0;">{esc(location)}</td></tr>
        <tr><td style="padding:4px 10px 4px 0;"><strong>Business Purpose:</strong></td><td style="padding:4px 0;">{esc(purpose)}</td></tr>
        <tr><td style="padding:4px 10px 4px 0;"><strong>Departure Date:</strong></td><td style="padding:4px 0;">{esc(departure_date)}</td></tr>
        <tr><td style="padding:4px 10px 4px 0;"><strong>Return Date:</strong></td><td style="padding:4px 0;">{esc(return_date)}</td></tr>
        <tr><td style="padding:4px 10px 4px 0;"><strong>Per Diem Total:</strong></td><td style="padding:4px 0;">${per_diem_total:,.2f}</td></tr>
        <tr><td style="padding:4px 10px 4px 0;"><strong>Total Spend:</strong></td><td style="padding:4px 0;">${total_spend:,.2f}</td></tr>
        <tr><td style="padding:4px 10px 4px 0;"><strong>Company Paid:</strong></td><td style="padding:4px 0;">${company_paid:,.2f}</td></tr>
        <tr><td style="padding:4px 10px 4px 0;"><strong>Employee Paid:</strong></td><td style="padding:4px 0;">${employee_paid:,.2f}</td></tr>
        <tr><td style="padding:4px 10px 4px 0;"><strong>Reimbursement Due:</strong></td><td style="padding:4px 0;">${reimbursement_due:,.2f}</td></tr>
      </table>

      {lines_html}

      <p>Please let me know if any additional information is required.</p>

      <p>Best regards,<br>{esc(employee_name)}</p>
    </div>
    """
    return html


def send_email_with_attachments(
    subject: str,
    html_body: str,
    employee_email: str,
    attachments: List[Dict[str, Any]],
) -> int:
    sg = SendGridAPIClient(sendgrid_api_key())

    msg = Mail(
        from_email=Email(secret_get("SENDER_EMAIL")),
        to_emails=To(secret_get("FINANCE_EMAIL")),
        subject=subject,
        html_content=html_body,
    )

    msg.add_cc(Cc(secret_get("APPROVER_EMAIL")))
    msg.add_cc(Cc(employee_email))

    for a in attachments:
        b = a["content_bytes"]
        encoded = base64.b64encode(b).decode("utf-8")
        msg.add_attachment(
            Attachment(
                FileContent(encoded),
                FileName(a["filename"]),
                FileType(a["mime_type"]),
                Disposition("attachment"),
            )
        )

    resp = sg.send(msg)
    return resp.status_code


def _each_date_inclusive(start_d: date, end_d: date):
    cur = start_d
    while cur <= end_d:
        yield cur
        cur = cur + timedelta(days=1)


def _normalize_sheet_name(name: str) -> str:
    parts = str(name or "").strip().split()
    return " ".join(parts).lower()


def _safe_float(val) -> float:
    try:
        return float(val or 0)
    except Exception:
        return 0.0


def _set_zero(ws, row: int, col: int) -> None:
    ws.cell(row=row, column=col).value = 0.0


def _add_amount_to_cell(ws, row: int, col: int, amount: float) -> None:
    amt = float(amount or 0)
    existing = ws.cell(row=row, column=col).value
    existing_float = _safe_float(existing)
    ws.cell(row=row, column=col).value = existing_float + amt


def _append_note(ws, row: int, note: str) -> None:
    if not note:
        return
    cell = ws.cell(row=row, column=NN_COL_NOTES)
    existing = cell.value or ""
    if str(existing).strip():
        cell.value = f"{existing}; {note}"
    else:
        cell.value = note


def _find_cell_by_exact_text(ws, target_text: str) -> Optional[Tuple[int, int]]:
    t = str(target_text).strip()
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            v = cell.value
            if v is None:
                continue
            if str(v).strip() == t:
                return (cell.row, cell.column)
    return None


def _write_killol_notes_values(ws, values: Dict[str, float]) -> None:
    for label, amount in values.items():
        pos = _find_cell_by_exact_text(ws, label)
        if not pos:
            continue
        r, c = pos
        if c <= 1:
            continue
        ws.cell(row=r, column=c - 1).value = float(amount)


def _find_template_last_row(ws) -> int:
    r = NN_START_ROW
    last = NN_START_ROW
    for _ in range(0, 60):
        v = ws.cell(row=r, column=NN_COL_DATE).value
        if v is None or str(v).strip() == "":
            break
        last = r
        r += 1
    return last


_CELL_REF_RE = re.compile(r"(\$?[A-Z]{1,3})(\$?)(\d+)")


def _rewrite_formula_row_refs(formula: str, src_row: int, dst_row: int) -> str:
    """
    Updates cell refs in a formula so any reference to src_row becomes dst_row.
    Examples:
      B6 -> B7
      $B6 -> $B7
      B$6 -> B$7
      $B$6 -> $B$7
    Only changes row numbers equal to src_row.
    """
    def repl(m):
        col = m.group(1)
        dollar_row = m.group(2)
        row_num = int(m.group(3))
        if row_num == src_row:
            return f"{col}{dollar_row}{dst_row}"
        return m.group(0)

    return _CELL_REF_RE.sub(repl, formula)


def _copy_formulas_only_with_row_fix(ws, src_row: int, dst_row: int, min_col: int, max_col: int) -> None:
    for c in range(min_col, max_col + 1):
        src_val = ws.cell(row=src_row, column=c).value
        if isinstance(src_val, str) and src_val.startswith("="):
            ws.cell(row=dst_row, column=c).value = _rewrite_formula_row_refs(src_val, src_row, dst_row)


def calc_nn_per_diem_for_day(d: date, departure_date: date, return_date: date) -> float:
    if d == departure_date or d == return_date:
        return float(NN_TRAVEL_DAY_RATE)
    return float(NN_NON_TRAVEL_DAY_RATE)


def calc_nn_per_diem_total(departure_date: date, return_date: date) -> float:
    if return_date < departure_date:
        return 0.0
    total = 0.0
    for d in _each_date_inclusive(departure_date, return_date):
        total += calc_nn_per_diem_for_day(d, departure_date, return_date)
    return total


def _clamp_expense_date_to_trip(expense_d: date, departure_date: date, return_date: date) -> date:
    if expense_d < departure_date:
        return departure_date
    if expense_d > return_date:
        return return_date
    return expense_d


def generate_nn_excel_bytes(
    employee_name: str,
    departure_date: date,
    return_date: date,
    expenses: List[Dict[str, Any]],
    sheet_name: str,
) -> bytes:
    template_path = os.path.join(TEMPLATES_DIR, NN_TEMPLATE_FILENAME)
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"NN template not found at {template_path}")

    wb = load_workbook(template_path)

    resolved_sheet = None
    if sheet_name in wb.sheetnames:
        resolved_sheet = sheet_name
    else:
        target_norm = _normalize_sheet_name(sheet_name)
        for s in wb.sheetnames:
            if _normalize_sheet_name(s) == target_norm:
                resolved_sheet = s
                break

    if not resolved_sheet:
        available = ", ".join(wb.sheetnames)
        raise ValueError(f"Sheet not found. Available sheets: {available}")

    ws = wb[resolved_sheet]

    ws["A1"].value = employee_name

    template_last_row = _find_template_last_row(ws)

    date_to_row: Dict[date, int] = {}
    r = NN_START_ROW

    for d in _each_date_inclusive(departure_date, return_date):
        if r > template_last_row:
            _copy_formulas_only_with_row_fix(ws, template_last_row, r, 1, NN_COL_NOTES)

        ws.cell(row=r, column=NN_COL_DATE).value = d
        ws.cell(row=r, column=NN_COL_INCIDENTALS).value = calc_nn_per_diem_for_day(d, departure_date, return_date)

        _set_zero(ws, r, NN_COL_AIRFARE)
        _set_zero(ws, r, NN_COL_LODGING)
        _set_zero(ws, r, NN_COL_VEHICLE_MILEAGE)
        _set_zero(ws, r, NN_COL_RENTAL_CAR)
        _set_zero(ws, r, NN_COL_MISC_TRAVEL)

        ws.cell(row=r, column=NN_COL_NOTES).value = ""

        date_to_row[d] = r
        r += 1

    category_to_col = {
        "Airfare": NN_COL_AIRFARE,
        "Hotel": NN_COL_LODGING,
        "Rental Car": NN_COL_RENTAL_CAR,
        "Gas for Rental Car": NN_COL_MISC_TRAVEL,
        "Airport Parking": NN_COL_MISC_TRAVEL,
        "Taxi or Uber to Airport": NN_COL_MISC_TRAVEL,
        "Other": NN_COL_MISC_TRAVEL,
    }

    company_paid_hotel = 0.0
    company_paid_total = 0.0
    employee_paid_total = 0.0
    company_paid_market_at_hotel = 0.0

    for e in expenses:
        d = e.get("expense_date")
        if not isinstance(d, date):
            continue

        # NN behavior: clamp expenses outside the trip window onto the nearest trip day
        d2 = _clamp_expense_date_to_trip(d, departure_date, return_date)

        row = date_to_row.get(d2)
        if row is None:
            continue

        cat = str(e.get("category") or "")
        col = category_to_col.get(cat, NN_COL_MISC_TRAVEL)

        amt = float(e.get("amount") or 0)
        if amt != 0:
            _add_amount_to_cell(ws, row, col, amt)

        paid_by = str(e.get("paid_by") or "").strip()
        desc = str(e.get("description") or "").strip()

        if paid_by == "Performa":
            company_paid_total += amt
            if cat == "Hotel":
                company_paid_hotel += amt
            if "market" in desc.lower():
                company_paid_market_at_hotel += amt
        else:
            employee_paid_total += amt

        if desc and paid_by:
            note = f"{cat}: {desc} (Paid by {paid_by})"
        elif desc:
            note = f"{cat}: {desc}"
        elif paid_by:
            note = f"{cat} (Paid by {paid_by})"
        else:
            note = f"{cat}"

        # If clamped, show original date in note
        if d2 != d:
            note = f"{note} (Original date {d})"

        _append_note(ws, row, note)

    nn_per_diem_total = calc_nn_per_diem_total(departure_date, return_date)
    total_expense_report = nn_per_diem_total + company_paid_total + employee_paid_total
    amount_due_to_employee = nn_per_diem_total + employee_paid_total
    amount_due_to_performa = company_paid_total

    killol_values = {
        "Total Expense Report": total_expense_report,
        "Amout Charged to Performa Card for Hotel": company_paid_hotel,
        "Amout for Hotel only on Performa Card": company_paid_hotel,
        "Amout for Market at hotel on Performa Card": company_paid_market_at_hotel,
        "Amount due to Brian*": amount_due_to_employee,
        "Amount due to Performa": amount_due_to_performa,
    }
    _write_killol_notes_values(ws, killol_values)

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


# -----------------------------
# App state
# -----------------------------
if "expenses" not in st.session_state:
    st.session_state.expenses = []

if "editing_expense_idx" not in st.session_state:
    st.session_state.editing_expense_idx = None

if st.session_state.pop("submit_success", False):
    st.success("Submitted successfully. Check your email for the package.")


def _render_email_setup_help(*, widget_key: str) -> None:
    st.markdown(
        "Paste the encoded line into **[Streamlit Cloud](https://share.streamlit.io) → your app → "
        "Settings → Secrets**, save, then **Reboot**."
    )
    raw_key_for_b64 = st.text_input(
        "Paste SendGrid API key to encode",
        type="password",
        placeholder="SG.xxxxxxxxxxxx",
        key=widget_key,
    )
    if raw_key_for_b64.strip():
        try:
            b64_value = make_sendgrid_b64(raw_key_for_b64)
            st.code(f'SENDGRID_API_KEY_B64 = "{b64_value}"', language="toml")
        except Exception as ex:
            st.error(f"Could not encode key: {ex}")

    diag = sendgrid_key_diagnostics()
    st.json(diag)
    if not diag.get("ok"):
        st.warning("Key not valid yet — use SENDGRID_API_KEY_B64 (one line in Secrets).")


with st.sidebar:
    st.subheader("SendGrid setup")
    _render_email_setup_help(widget_key="sendgrid_key_encode_sidebar")


# -----------------------------
# UI
# -----------------------------
st.markdown(
    '<p style="font-size:1.75rem;font-weight:700;letter-spacing:0.06em;margin:0 0 0.15rem 0;">PERFORMA</p>',
    unsafe_allow_html=True,
)
st.title("Performa Expense Report")
st.caption("© 2026 bkw")
st.caption(
    "Guam mode generates Excel plus receipts. NN mode populates the NN Excel template plus receipts. "
    "Emails Finance, Approver, and Employee."
)

with st.expander("Email setup help (same as sidebar → SendGrid setup)", expanded=not has_sendgrid_key()):
    _render_email_setup_help(widget_key="sendgrid_key_encode_main")

st.subheader("Report Type")
report_type = st.radio(
    "Select report type",
    ["Guam", "NN (Navajo Nation)"],
    horizontal=True,
)

# NN sheet chooser
nn_sheet_choice = NN_SHEET_NAME_DEFAULT
if report_type == "NN (Navajo Nation)":
    template_path_ui = os.path.join(TEMPLATES_DIR, NN_TEMPLATE_FILENAME)
    if not os.path.exists(template_path_ui):
        st.error(f"NN template file not found at {template_path_ui}")
    else:
        try:
            wb_ui = load_workbook(template_path_ui, read_only=True)
            sheetnames = wb_ui.sheetnames
            wb_ui.close()

            default_idx = 0
            target_norm = _normalize_sheet_name(NN_SHEET_NAME_DEFAULT)
            for i, s in enumerate(sheetnames):
                if _normalize_sheet_name(s) == target_norm:
                    default_idx = i
                    break
            nn_sheet_choice = st.selectbox("NN Sheet", sheetnames, index=default_idx)
        except Exception as ex:
            st.error(f"Unable to read NN template sheets: {ex}")

st.subheader("Trip Information")

col1, col2 = st.columns(2)
with col1:
    employee_name = st.text_input("Employee Name")
    employee_email = st.text_input("Employee Email")
with col2:
    location = st.text_input("Trip Location")
    purpose = st.text_area("Business Purpose", height=80)

col3, col4 = st.columns(2)
with col3:
    departure_date = st.date_input("Departure Date", value=date.today())
with col4:
    return_date = st.date_input("Return Date", value=date.today())

trip_days = calc_trip_days(departure_date, return_date)

if report_type == "Guam":
    per_diem_total_display = calculate_per_diem_total(
        trip_days,
        first_last_rate=PER_DIEM_FIRST_LAST_RATE,
        middle_rate=PER_DIEM_MIDDLE_RATE,
    )
    st.info(
        format_per_diem_summary(
            trip_days,
            first_last_rate=PER_DIEM_FIRST_LAST_RATE,
            middle_rate=PER_DIEM_MIDDLE_RATE,
        )
    )
else:
    per_diem_total_display = calc_nn_per_diem_total(departure_date, return_date)
    st.info(
        f"NN per diem is ${NN_TRAVEL_DAY_RATE:,.0f} on travel days and ${NN_NON_TRAVEL_DAY_RATE:,.0f} on non travel days, "
        f"{trip_days} day(s), total ${per_diem_total_display:,.2f}"
    )

st.subheader("Expenses")

with st.expander("Add an expense", expanded=True):
    with st.form("add_expense_form", clear_on_submit=True):
        c1, c2, c3 = st.columns([2, 2, 2])
        with c1:
            category = st.selectbox("Category", CATEGORIES)
        with c2:
            expense_date = st.date_input("Expense Date", value=departure_date)
        with c3:
            paid_by = st.radio("Paid By", ["Employee", "Performa"], horizontal=True)

        description = st.text_input("Description (optional)")
        amount = st.number_input("Amount", min_value=0.0, value=0.0, step=1.0, format="%.2f")

        receipt_file = st.file_uploader(
            "Receipt (optional)",
            type=["pdf", "png", "jpg", "jpeg"],
            accept_multiple_files=False,
            help="Accepted: PDF, JPG, JPEG, PNG",
        )

        add_btn = st.form_submit_button("Add Expense")

        if add_btn:
            st.session_state.expenses.append(
                {
                    "category": category,
                    "expense_date": expense_date,
                    "paid_by": paid_by,
                    "description": description,
                    "amount": float(amount),
                    "receipt_file": receipt_file,
                }
            )
            st.success("Expense added.")


st.subheader("Summary")

totals = calc_totals(st.session_state.expenses)
total_spend = totals["total_spend"]
company_paid = totals["company_paid"]
employee_paid = totals["employee_paid"]

reimbursement_due = per_diem_total_display + employee_paid

s1, s2, s3, s4 = st.columns(4)
s1.metric("Total Spend", f"${total_spend:,.2f}")
s2.metric("Company Paid", f"${company_paid:,.2f}")
s3.metric("Employee Paid", f"${employee_paid:,.2f}")
s4.metric("Employee Due", f"${reimbursement_due:,.2f}")

def _expense_line_label(idx: int, e: Dict[str, Any]) -> str:
    receipt_note = "receipt" if e.get("receipt_file") else "no receipt"
    desc = (e.get("description") or "").strip() or "—"
    return (
        f"{idx + 1}. {e['category']} on {e['expense_date']}, {desc}, "
        f"${float(e['amount']):,.2f}, {e['paid_by']}, {receipt_note}"
    )


st.subheader("Current Line Items")
if not st.session_state.expenses:
    st.write("No expenses added yet.")
    st.session_state.editing_expense_idx = None
else:
    for idx, e in enumerate(st.session_state.expenses):
        st.write(_expense_line_label(idx, e))

    line_options = list(range(len(st.session_state.expenses)))
    selected_line = st.selectbox(
        "Select line item to edit or remove",
        options=line_options,
        format_func=lambda i: _expense_line_label(i, st.session_state.expenses[i]),
        key="selected_expense_line",
    )

    action_col1, action_col2 = st.columns(2)
    with action_col1:
        if st.button("Edit Selected", use_container_width=True):
            st.session_state.editing_expense_idx = selected_line
            st.rerun()
    with action_col2:
        if st.button("Remove Selected", use_container_width=True):
            st.session_state.expenses.pop(selected_line)
            if st.session_state.editing_expense_idx == selected_line:
                st.session_state.editing_expense_idx = None
            elif (
                st.session_state.editing_expense_idx is not None
                and st.session_state.editing_expense_idx > selected_line
            ):
                st.session_state.editing_expense_idx -= 1
            st.rerun()

    edit_idx = st.session_state.editing_expense_idx
    if edit_idx is not None and edit_idx < len(st.session_state.expenses):
        expense = st.session_state.expenses[edit_idx]
        with st.expander(f"Edit line item {edit_idx + 1}", expanded=True):
            with st.form("edit_expense_form"):
                ec1, ec2, ec3 = st.columns([2, 2, 2])
                with ec1:
                    cat_index = (
                        CATEGORIES.index(expense["category"])
                        if expense["category"] in CATEGORIES
                        else 0
                    )
                    edit_category = st.selectbox(
                        "Category",
                        CATEGORIES,
                        index=cat_index,
                    )
                with ec2:
                    edit_date = st.date_input(
                        "Expense Date",
                        value=expense["expense_date"],
                    )
                with ec3:
                    edit_paid_by = st.radio(
                        "Paid By",
                        ["Employee", "Performa"],
                        index=0 if expense["paid_by"] == "Employee" else 1,
                        horizontal=True,
                    )

                edit_description = st.text_input(
                    "Description (optional)",
                    value=expense.get("description") or "",
                )
                edit_amount = st.number_input(
                    "Amount",
                    min_value=0.0,
                    value=float(expense.get("amount", 0)),
                    step=1.0,
                    format="%.2f",
                )

                if expense.get("receipt_file"):
                    st.caption(f"Current receipt: {expense['receipt_file'].name}")
                edit_receipt = st.file_uploader(
                    "Receipt (optional, upload to replace)",
                    type=["pdf", "png", "jpg", "jpeg"],
                    accept_multiple_files=False,
                    help="Leave empty to keep the existing receipt.",
                )

                save_col, cancel_col = st.columns(2)
                with save_col:
                    save_edit = st.form_submit_button("Save Changes", use_container_width=True)
                with cancel_col:
                    cancel_edit = st.form_submit_button("Cancel", use_container_width=True)

                if save_edit:
                    st.session_state.expenses[edit_idx] = {
                        "category": edit_category,
                        "expense_date": edit_date,
                        "paid_by": edit_paid_by,
                        "description": edit_description,
                        "amount": float(edit_amount),
                        "receipt_file": edit_receipt or expense.get("receipt_file"),
                    }
                    st.session_state.editing_expense_idx = None
                    st.rerun()

                if cancel_edit:
                    st.session_state.editing_expense_idx = None
                    st.rerun()
    elif edit_idx is not None:
        st.session_state.editing_expense_idx = None


st.divider()
st.caption(
    f"Attachment limit enforced at {MAX_ATTACHMENT_MB:,.0f} MB total for receipts plus the report file."
)

submit = st.button("Submit Expense Report", type="primary")

if submit:
    missing = []
    if not employee_name.strip():
        missing.append("Employee Name")
    if not employee_email.strip():
        missing.append("Employee Email")
    if not location.strip():
        missing.append("Trip Location")
    if not purpose.strip():
        missing.append("Business Purpose")

    if return_date < departure_date:
        missing.append("Return Date must be on or after Departure Date")

    if report_type == "NN (Navajo Nation)":
        template_path_check = os.path.join(TEMPLATES_DIR, NN_TEMPLATE_FILENAME)
        if not os.path.exists(template_path_check):
            missing.append(f"NN template file missing: {template_path_check}")

    if missing:
        st.error("Please complete the following fields: " + ", ".join(missing))
        st.stop()

    unset_secrets = missing_email_secrets()
    if unset_secrets:
        st.error(
            "Email is not configured. Add these keys to `.streamlit/secrets.toml` "
            f"(see `.streamlit/secrets.toml.example`): {', '.join(unset_secrets)}"
        )
        st.stop()

    key_error = verify_sendgrid_key_with_api()
    if key_error:
        st.error(key_error)
        st.stop()

    # Guam generator inputs, unchanged
    trip_info = {
        "employee_name": employee_name,
        "employee_email": employee_email,
        "location": location,
        "purpose": purpose,
        "departure_date": departure_date,
        "return_date": return_date,
        "trip_days": trip_days,
        "per_diem_first_last_rate": PER_DIEM_FIRST_LAST_RATE,
        "per_diem_middle_rate": PER_DIEM_MIDDLE_RATE,
        "per_diem_total": per_diem_total_display,
        "total_spend": total_spend,
        "company_paid": company_paid,
        "employee_paid": employee_paid,
        "reimbursement_due": per_diem_total_display + employee_paid,
    }

    attachments: List[Dict[str, Any]] = []

    if report_type == "Guam":
        excel_bytes = generate_excel(trip_info, st.session_state.expenses)
        attachments.append(
            {
                "filename": f"Expense_Report_{employee_name.replace(' ', '_')}.xlsx",
                "content_bytes": excel_bytes,
                "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            }
        )
    else:
        try:
            nn_bytes = generate_nn_excel_bytes(
                employee_name=employee_name,
                departure_date=departure_date,
                return_date=return_date,
                expenses=st.session_state.expenses,
                sheet_name=nn_sheet_choice,
            )
        except Exception as ex:
            st.error(f"NN report generation failed: {ex}")
            st.stop()

        attachments.append(
            {
                "filename": f"NN_Expense_Report_{employee_name.replace(' ', '_')}.xlsx",
                "content_bytes": nn_bytes,
                "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            }
        )

    for i, e in enumerate(st.session_state.expenses, start=1):
        f = e.get("receipt_file")
        if f is None:
            continue
        b = bytes_from_uploaded_file(f)
        ext = (f.name.split(".")[-1] or "").lower()
        if ext in ["jpg", "jpeg"]:
            mime = "image/jpeg"
        elif ext == "png":
            mime = "image/png"
        else:
            mime = "application/pdf"

        safe_cat = str(e.get("category", "Receipt")).replace(" ", "_")
        filename = f"{i:02d}_{safe_cat}_{employee_name.replace(' ', '_')}.{ext if ext else 'pdf'}"
        attachments.append({"filename": filename, "content_bytes": b, "mime_type": mime})

    total_bytes = sum(len(a["content_bytes"]) for a in attachments)
    max_bytes = int(MAX_ATTACHMENT_MB * 1024 * 1024)
    if total_bytes > max_bytes:
        st.error(
            f"Attachments are too large: {total_bytes/1024/1024:,.2f} MB. "
            f"Limit is {MAX_ATTACHMENT_MB:,.0f} MB. Remove some receipts or compress them."
        )
        st.stop()

    subject = f"Expense Report Submitted, {employee_name}, {location}, {departure_date} to {return_date}"
    if report_type == "NN (Navajo Nation)":
        subject = f"NN Expense Report Submitted, {employee_name}, {location}, {departure_date} to {return_date}"

    html_body = build_email_html(
        employee_name=employee_name,
        employee_email=employee_email,
        location=location,
        purpose=purpose,
        departure_date=departure_date,
        return_date=return_date,
        per_diem_total=per_diem_total_display,
        total_spend=total_spend,
        company_paid=company_paid,
        employee_paid=employee_paid,
        reimbursement_due=reimbursement_due,
        expenses=st.session_state.expenses,
    )

    try:
        status_code = send_email_with_attachments(
            subject=subject,
            html_body=html_body,
            employee_email=employee_email,
            attachments=attachments,
        )

        if 200 <= int(status_code) < 300:
            st.session_state.expenses = []
            st.session_state.editing_expense_idx = None
            st.session_state.submit_success = True
            st.rerun()
        else:
            st.error(f"SendGrid returned status code: {status_code}")
    except Exception as ex:
        st.error(format_sendgrid_error(ex))
